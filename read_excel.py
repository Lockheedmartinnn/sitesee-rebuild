import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('/home/ubuntu/upload/RevenueaccelerationNPV_FisherEquationVersion_v1.00.xlsx', data_only=True)

# Print sheet names
print("Sheets:", wb.sheetnames)

# Read the main sheet (assuming it's the first one or named appropriately)
sheet = wb.active
print(f"Active Sheet: {sheet.title}")

# Print key inputs and outputs to verify structure
print("\n--- Inputs ---")
print(f"WACC (B2): {sheet['B2'].value}")
print(f"Annual Rent (B3): {sheet['B3'].value}")
print(f"Annual Rent Escalator (B4): {sheet['B4'].value}")
print(f"Annual Inflation (B5): {sheet['B5'].value}")
print(f"Lease Term (B6): {sheet['B6'].value}")
print(f"Rent Commencement Acceleration (B7): {sheet['B7'].value}")
print(f"Rent Payment Frequency (B8): {sheet['B8'].value}")
print(f"Tax Rate (B9): {sheet['B9'].value}")
print(f"Operating Margin (B10): {sheet['B10'].value}")
print(f"Enterprise Valuation EBITDA Multiple (B11): {sheet['B11'].value}")

print("\n--- Outputs ---")
print(f"NPV (B14): {sheet['B14'].value}")
print(f"EV Uplift (B15): {sheet['B15'].value}")

# Read the calculation table (assuming it starts around row 13/14 based on previous screenshots)
# Let's scan rows 13-25 to find the table headers and data
print("\n--- Calculation Table (Rows 13-25) ---")
for row in sheet.iter_rows(min_row=13, max_row=25, values_only=True):
    print(row)
