import openpyxl
import pandas as pd

def analyze_ev_uplift():
    print("\n=== Analyzing EV Uplift (RevenueaccelerationNPV_FisherEquationVersion_v1.00.xlsx) ===")
    try:
        wb = openpyxl.load_workbook('/home/ubuntu/upload/RevenueaccelerationNPV_FisherEquationVersion_v1.00.xlsx', data_only=True)
        sheet = wb.active
        
        # Inputs
        inputs = {
            'WACC': sheet['B2'].value,
            'Annual Rent': sheet['B3'].value,
            'Escalator': sheet['B4'].value,
            'Inflation': sheet['B5'].value,
            'Term': sheet['B6'].value,
            'Days': sheet['B7'].value,
            'Freq': sheet['B8'].value,
            'Tax': sheet['B9'].value,
            'Margin': sheet['B10'].value,
            'Multiple': sheet['B11'].value
        }
        print("Inputs:", inputs)
        
        # Outputs
        outputs = {
            'NPV': sheet['B14'].value,
            'EV Uplift': sheet['B15'].value
        }
        print("Outputs:", outputs)
        
        # Table Logic (Rows 13-25)
        print("\nTable Sample (First 5 rows):")
        for row in sheet.iter_rows(min_row=13, max_row=18, values_only=True):
            print(row)
            
    except Exception as e:
        print(f"Error reading EV Uplift file: {e}")

def analyze_nexdt():
    print("\n=== Analyzing NexDT Value Calculator (NexDT_Value_Calculator-7.xlsx) ===")
    try:
        wb = openpyxl.load_workbook('/home/ubuntu/upload/NexDT_Value_Calculator-7.xlsx', data_only=True)
        sheet = wb.active # Assuming first sheet is the calculator
        
        # Scan for inputs based on labels (since layout might differ)
        # Printing first 20 rows to understand structure
        print("\nFirst 20 rows of NexDT sheet:")
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True)):
            print(f"Row {i+1}: {row}")
            
    except Exception as e:
        print(f"Error reading NexDT file: {e}")

if __name__ == "__main__":
    analyze_ev_uplift()
    analyze_nexdt()
