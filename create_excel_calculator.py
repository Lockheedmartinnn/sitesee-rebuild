import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create a new workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "EBITDA & EV Calculator"

# Define styles
header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
subheader_font = Font(name='Calibri', size=12, bold=True, color='1F4E78')
input_font = Font(name='Calibri', size=11, color='000000')
input_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid') # Light green for inputs
calc_font = Font(name='Calibri', size=11, bold=True, color='000000')
result_font = Font(name='Calibri', size=14, bold=True, color='000000')
result_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid') # Yellow for results
border_thin = Side(border_style="thin", color="000000")
border_medium = Side(border_style="medium", color="000000")
border_all = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)

# --- INPUTS SECTION ---
ws['B2'] = "EBITDA & ENTERPRISE VALUE CALCULATOR"
ws['B2'].font = Font(name='Calibri', size=18, bold=True, color='1F4E78')

ws['B4'] = "INPUTS"
ws['B4'].font = header_font
ws['B4'].fill = header_fill
ws['C4'].fill = header_fill # Extend fill

inputs = [
    ("Number of Sites", 1, "C5"),
    ("Annual Rent per Site", 25000, "C6"),
    ("Operating Margin (%)", 0.80, "C7"),
    ("Acceleration Days", 90, "C8"),
    ("EBITDA Multiple", 15, "C9"),
    ("Operational Savings (Annual per Site)", 800, "C10"),
    ("FTR Improvement (Annual per Site)", 1600, "C11")
]

row = 5
for label, value, cell_ref in inputs:
    ws[f'B{row}'] = label
    ws[f'B{row}'].font = input_font
    ws[f'B{row}'].border = border_all
    
    ws[f'C{row}'] = value
    ws[f'C{row}'].font = input_font
    ws[f'C{row}'].fill = input_fill
    ws[f'C{row}'].border = border_all
    ws[f'C{row}'].alignment = Alignment(horizontal='center')
    
    if "Rent" in label or "Savings" in label or "Improvement" in label:
        ws[f'C{row}'].number_format = '$#,##0'
    elif "Margin" in label:
        ws[f'C{row}'].number_format = '0.00%'
    
    row += 1

# --- CALCULATIONS SECTION ---
row += 2
ws[f'B{row}'] = "EBITDA CALCULATIONS"
ws[f'B{row}'].font = header_font
ws[f'B{row}'].fill = header_fill
ws[f'C{row}'].fill = header_fill

row += 1
# EBITDA from Revenue Acceleration
ws[f'B{row}'] = "EBITDA from Revenue Acceleration"
ws[f'B{row}'].font = subheader_font
ws[f'B{row}'].border = border_all

# Formula: (Annual Rent / 365) * Acceleration Days * Margin * Num Sites
ws[f'C{row}'] = "= (C6 / 365) * C8 * C7 * C5"
ws[f'C{row}'].font = calc_font
ws[f'C{row}'].number_format = '$#,##0.00'
ws[f'C{row}'].border = border_all
ws[f'C{row}'].alignment = Alignment(horizontal='right')

row += 1
# EBITDA from Operational Savings
ws[f'B{row}'] = "EBITDA from Operational Savings"
ws[f'B{row}'].font = subheader_font
ws[f'B{row}'].border = border_all

# Formula: Operational Savings * Num Sites
ws[f'C{row}'] = "= C10 * C5"
ws[f'C{row}'].font = calc_font
ws[f'C{row}'].number_format = '$#,##0.00'
ws[f'C{row}'].border = border_all
ws[f'C{row}'].alignment = Alignment(horizontal='right')

row += 1
# EBITDA from FTR Improvement
ws[f'B{row}'] = "EBITDA from FTR Improvement"
ws[f'B{row}'].font = subheader_font
ws[f'B{row}'].border = border_all

# Formula: FTR Improvement * Num Sites
ws[f'C{row}'] = "= C11 * C5"
ws[f'C{row}'].font = calc_font
ws[f'C{row}'].number_format = '$#,##0.00'
ws[f'C{row}'].border = border_all
ws[f'C{row}'].alignment = Alignment(horizontal='right')

row += 1
# Total Annual EBITDA
ws[f'B{row}'] = "Total Annual EBITDA"
ws[f'B{row}'].font = result_font
ws[f'B{row}'].fill = result_fill
ws[f'B{row}'].border = border_all

ws[f'C{row}'] = f"= SUM(C{row-3}:C{row-1})"
ws[f'C{row}'].font = result_font
ws[f'C{row}'].fill = result_fill
ws[f'C{row}'].number_format = '$#,##0.00'
ws[f'C{row}'].border = border_all
ws[f'C{row}'].alignment = Alignment(horizontal='right')

# --- ENTERPRISE VALUE SECTION ---
row += 2
ws[f'B{row}'] = "ENTERPRISE VALUE"
ws[f'B{row}'].font = header_font
ws[f'B{row}'].fill = header_fill
ws[f'C{row}'].fill = header_fill

row += 1
ws[f'B{row}'] = "Enterprise Value Uplift"
ws[f'B{row}'].font = result_font
ws[f'B{row}'].fill = result_fill
ws[f'B{row}'].border = border_all

# Formula: Total EBITDA * Multiple
ws[f'C{row}'] = f"= C{row-3} * C9"
ws[f'C{row}'].font = result_font
ws[f'C{row}'].fill = result_fill
ws[f'C{row}'].number_format = '$#,##0.00'
ws[f'C{row}'].border = border_all
ws[f'C{row}'].alignment = Alignment(horizontal='right')

# Adjust column widths
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 25

# Save the workbook
wb.save("/home/ubuntu/sitesee-rebuild/EBITDA_EV_Calculator.xlsx")
print("Excel file created successfully: /home/ubuntu/sitesee-rebuild/EBITDA_EV_Calculator.xlsx")
