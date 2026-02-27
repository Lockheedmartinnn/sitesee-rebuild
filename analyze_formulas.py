import openpyxl

def analyze_formulas():
    print("\n=== Analyzing Formulas in EV Uplift (RevenueaccelerationNPV_FisherEquationVersion_v1.00.xlsx) ===")
    try:
        wb = openpyxl.load_workbook('/home/ubuntu/upload/RevenueaccelerationNPV_FisherEquationVersion_v1.00.xlsx', data_only=False) # data_only=False to get formulas
        sheet = wb.active
        
        # Check where B5 (Inflation) and B9 (Tax) are used
        # We suspect they are used in the calculation table (Rows 13-25) or in intermediate cells
        
        print("Checking formulas in Rows 13-25:")
        for row in sheet.iter_rows(min_row=13, max_row=25):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and ('B5' in cell.value or 'B9' in cell.value):
                    print(f"Cell {cell.coordinate}: {cell.value}")
                    
        # Also check B14 (NPV) and B15 (EV Uplift) directly
        print(f"NPV (B14): {sheet['B14'].value}")
        print(f"EV Uplift (B15): {sheet['B15'].value}")

        # Check if there are hidden calculation columns (e.g., C, D, E...)
        print("\nChecking hidden columns (C-Z) in Row 15 (just as a sample):")
        for col in range(3, 27): # C to Z
            cell = sheet.cell(row=15, column=col)
            if cell.value:
                print(f"Cell {cell.coordinate}: {cell.value}")

    except Exception as e:
        print(f"Error reading EV Uplift file: {e}")

if __name__ == "__main__":
    analyze_formulas()
